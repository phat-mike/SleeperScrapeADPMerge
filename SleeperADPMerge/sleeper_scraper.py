import requests
import pandas as pd
import pickle
import os
from datetime import datetime
import time
from pathlib import Path
import sys

class SleeperAPIExporter:
    def __init__(self, cache_dir="sleeper_cache"):
        self.base_url = "https://api.sleeper.app/v1"
        self.cache_dir = Path(cache_dir)
        self.cache_dir.mkdir(exist_ok=True)
        self.cache_file = self.cache_dir / "sleeper_data.pkl"
    
    def _make_api_request(self, endpoint, description):
        """Make a single API request"""
        print(f"🌐 Fetching {description}...")
        try:
            response = requests.get(f"{self.base_url}/{endpoint}")
            response.raise_for_status()
            time.sleep(0.5)  # Be nice to the API
            return response.json()
        except requests.exceptions.RequestException as e:
            print(f"❌ Error fetching {description}: {e}")
            raise  # Re-raise to let caller handle it
    
    def validate_league_id(self, league_id):
        """Test if a league ID is valid by making a simple API call"""
        try:
            self._make_api_request(f"league/{league_id}", f"league {league_id} validation")
            return True
        except requests.exceptions.RequestException:
            return False
    
    def fetch_all_data(self, league_id=None, include_players=False, weeks=None):
        """Fetch all data from API and return as dict"""
        print("=== Fetching all data from Sleeper API ===")
        all_data = {}
        
        # Fetch players data if requested
        if include_players:
            all_data['players'] = self._make_api_request("players/nfl", "NFL players")
        
        # Fetch league data if league_id provided
        if league_id:
            all_data['league_info'] = self._make_api_request(f"league/{league_id}", "league info")
            all_data['users'] = self._make_api_request(f"league/{league_id}/users", "league users")
            all_data['rosters'] = self._make_api_request(f"league/{league_id}/rosters", "league rosters")
            
            # Fetch matchups for specific weeks
            if weeks:
                all_data['matchups'] = {}
                for week in weeks:
                    matchup_data = self._make_api_request(
                        f"league/{league_id}/matchups/{week}", 
                        f"week {week} matchups"
                    )
                    if matchup_data:
                        all_data['matchups'][week] = matchup_data
        
        # Add metadata
        all_data['_metadata'] = {
            'fetched_at': datetime.now().isoformat(),
            'league_id': league_id,
            'included_players': include_players,
            'weeks': weeks or []
        }
        
        return all_data
    
    def save_to_cache(self, data):
        """Save all data to cache file"""
        try:
            with open(self.cache_file, 'wb') as f:
                pickle.dump(data, f)
            print(f"✓ Saved all data to cache: {self.cache_file}")
        except (pickle.PickleError, IOError) as e:
            print(f"⚠️ Failed to save cache: {e}")
    
    def load_from_cache(self):
        """Load all data from cache file"""
        if not self.cache_file.exists():
            print("No cache file found")
            return None
        
        try:
            with open(self.cache_file, 'rb') as f:
                data = pickle.load(f)
            
            # Show cache info
            metadata = data.get('_metadata', {})
            fetched_at = metadata.get('fetched_at', 'Unknown')
            print(f"✓ Loaded cached data (fetched: {fetched_at})")
            
            return data
        except (pickle.PickleError, IOError) as e:
            print(f"❌ Failed to load cache: {e}")
            return None
    
    def get_data(self, league_id=None, include_players=False, weeks=None, use_cache=True):
        """Get data either from cache or fresh from API with automatic fallback"""
        
        if use_cache:
            cached_data = self.load_from_cache()
            if cached_data is not None:
                return cached_data
            else:
                print("⚠️ Cache requested but not found, falling back to API...")
        
        # Fetch fresh data from API
        fresh_data = self.fetch_all_data(league_id, include_players, weeks)
        
        # Save to cache
        self.save_to_cache(fresh_data)
        
        return fresh_data
    
    def clean_status_value(self, status):
        """Clean and standardize player status values"""
        if not status:
            return ''
        
        # Convert to string and clean up
        status_str = str(status).lower().strip()
        
        # Status mappings (case-insensitive)
        status_mappings = {
            'injured reserve': 'IR',
            'physically unable to perform': 'PUP',
            'practice squad': 'PS',
            # Add any other mappings you discover
            'non-football injury': 'NFI',
            'suspended': 'SUSP',
            'commissioner exempt': 'EXEMPT'
        }
        
        # Apply mappings
        for original, abbreviation in status_mappings.items():
            if status_str == original:
                return abbreviation
        
        # Return original if no mapping found
        return status_str

    def process_players_data(self, players_data, league_data=None):
        """Convert players dict to DataFrame with smart filtering"""
        if not players_data:
            return pd.DataFrame()
        
        print("Processing players data with smart filtering...")
        print(f"🔍 Starting with {len(players_data)} total players...")
        
        # Get valid fantasy positions from league data
        valid_fantasy_positions = set()
        if league_data and 'roster_positions' in league_data:
            valid_fantasy_positions = set(league_data['roster_positions'])
            print(f"🔍 Valid fantasy positions from league: {sorted(valid_fantasy_positions)}")
        else:
            # Fallback to common fantasy positions if no league data
            valid_fantasy_positions = {'QB', 'RB', 'WR', 'TE', 'K', 'DEF'}
            print(f"⚠️ No league roster positions found, using default: {sorted(valid_fantasy_positions)}")
        
        players_list = []
        processed_count = 0
        
        # Tracking filtered players
        filter_stats = {
            'inactive_status': 0,
            'duplicate_name': 0,
            'no_fantasy_position': 0,
            'invalid_fantasy_position': 0,
            'data_errors': 0,
            'kept': 0
        }
        
        for player_id, player_info in players_data.items():
            processed_count += 1
            if processed_count % 1000 == 0:
                print(f"🔍 Processed {processed_count} players...")
            
            try:
                if not isinstance(player_info, dict):
                    filter_stats['data_errors'] += 1
                    continue
                
                # Safe string extraction with None handling
                def safe_string(value, default=''):
                    """Safely convert value to string, handling None values"""
                    if value is None:
                        return default
                    return str(value)
                
                # Filter 1: Remove inactive players (with None handling)
                status = safe_string(player_info.get('status', '')).lower()
                if status == 'inactive':
                    filter_stats['inactive_status'] += 1
                    continue
                
                # Filter 2: Remove players with "duplicate" in name (with None handling)
                full_name = safe_string(player_info.get('full_name', '')).lower()
                first_name = safe_string(player_info.get('first_name', '')).lower()
                last_name = safe_string(player_info.get('last_name', '')).lower()
                
                if ('duplicate' in full_name or 
                    'duplicate' in first_name or 
                    'duplicate' in last_name):
                    filter_stats['duplicate_name'] += 1
                    continue
                
                # Filter 3: Only keep players with valid fantasy positions
                fantasy_positions = player_info.get('fantasy_positions', [])
                
                # Handle case where fantasy_positions might not be a list or is None
                if not fantasy_positions:
                    filter_stats['no_fantasy_position'] += 1
                    continue
                
                if not isinstance(fantasy_positions, (list, tuple)):
                    filter_stats['invalid_fantasy_position'] += 1
                    continue
                
                # Check if player has at least one valid fantasy position
                player_positions = set(fantasy_positions)
                if not player_positions.intersection(valid_fantasy_positions):
                    filter_stats['invalid_fantasy_position'] += 1
                    continue
                
                # Player passed all filters - include them
                filter_stats['kept'] += 1
                
                # Safe handling of fantasy_positions for display
                fantasy_positions_str = ', '.join(fantasy_positions)
                
                # Process numeric fields properly
                numeric_player_id = self.safe_numeric(player_id)
                height_inches = self.parse_height(player_info.get('height'))
                weight_lbs = self.safe_numeric(player_info.get('weight'))
                age_years = self.safe_numeric(player_info.get('age'))
                years_experience = self.safe_numeric(player_info.get('years_exp'))

                raw_status = player_info.get('status', '')
                cleaned_status = self.clean_status_value(raw_status)
                
                players_list.append({
                    'player_id': numeric_player_id,  # Now numeric
                    'full_name': safe_string(player_info.get('full_name', '')),
                    'first_name': safe_string(player_info.get('first_name', '')),
                    'last_name': safe_string(player_info.get('last_name', '')),
                    'position': safe_string(player_info.get('position', '')),
                    'team': safe_string(player_info.get('team', '')),
                    'age': age_years,  # Now numeric
                    'height_inches': height_inches,  # Now numeric (in inches)
                    'weight_lbs': weight_lbs,  # Now numeric
                    'years_exp': years_experience,  # Now numeric
                    'college': safe_string(player_info.get('college', '')),
                    'status': cleaned_status,
                    'active': player_info.get('active', ''),
                    'fantasy_positions': fantasy_positions_str
                })
                
            except Exception as e:
                # Handle individual player errors without stopping the whole process
                filter_stats['data_errors'] += 1
                print(f"⚠️ Skipping player {processed_count} (ID: {player_id}) due to error: {e}")
                continue  # Skip this player and continue with the next one
        
        # Print filtering summary
        print(f"\n🔍 Smart filtering results:")
        print(f"  📊 Total processed: {processed_count:,}")
        print(f"  ❌ Filtered out:")
        print(f"     - Inactive status: {filter_stats['inactive_status']:,}")
        print(f"     - Duplicate names: {filter_stats['duplicate_name']:,}")
        print(f"     - No fantasy positions: {filter_stats['no_fantasy_position']:,}")
        print(f"     - Invalid fantasy positions: {filter_stats['invalid_fantasy_position']:,}")
        print(f"     - Data errors: {filter_stats['data_errors']:,}")
        print(f"  ✅ Fantasy-relevant players kept: {filter_stats['kept']:,}")
        
        if processed_count > 0:
            print(f"  📉 Reduction: {((processed_count - filter_stats['kept']) / processed_count * 100):.1f}%")
        
        df = pd.DataFrame(players_list)
        
        # Ensure proper data types for Excel
        if not df.empty:
            # Convert specific columns to proper numeric types
            numeric_columns = ['player_id', 'age', 'height_inches', 'weight_lbs', 'years_exp']
            for col in numeric_columns:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors='coerce')
        
        print(f"🔍 DataFrame created successfully: {len(df)} rows, {len(df.columns)} columns")
        return df

    def process_rosters_data(self, rosters_data, users_data=None):
        """Convert rosters data to DataFrame"""
        if not rosters_data:
            return pd.DataFrame()
        
        print("Processing rosters data...")
        
        # Create user lookup
        user_lookup = {}
        if users_data:
            for user in users_data:
                user_lookup[user.get('user_id')] = user.get('display_name', 'Unknown')
        
        rosters_list = []
        try:
            for roster in rosters_data:
                settings = roster.get('settings', {})
                
                # Safe handling of players list
                players = roster.get('players', [])
                if players and isinstance(players, (list, tuple)):
                    players_str = ', '.join(players)
                else:
                    players_str = str(players) if players else ''
                
                rosters_list.append({
                    'roster_id': roster.get('roster_id'),
                    'owner_id': roster.get('owner_id'),
                    'owner_name': user_lookup.get(roster.get('owner_id'), 'Unknown'),
                    'wins': settings.get('wins', 0),
                    'losses': settings.get('losses', 0),
                    'ties': settings.get('ties', 0),
                    'fpts': settings.get('fpts', 0),
                    'fpts_against': settings.get('fpts_against', 0),
                    'total_moves': settings.get('total_moves', 0),
                    'waiver_position': settings.get('waiver_position', 0),
                    'players': players_str
                })
            
            print(f"🔍 Successfully processed {len(rosters_list)} rosters")
            return pd.DataFrame(rosters_list)
            
        except Exception as e:
            print(f"❌ Error processing rosters data: {e}")
            # Return partial data if possible
            if rosters_list:
                return pd.DataFrame(rosters_list)
            else:
                return pd.DataFrame()

    def add_excel_filters(self, filename, freeze_panes=True, bold_headers=True):
        """Add auto-filters and formatting to Excel worksheets"""
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font
            
            print("🔧 Adding Excel filters and formatting...")
            
            # Load the workbook
            workbook = load_workbook(filename)
            
            # Sheets that should have filters
            filterable_sheets = ['Players', 'Users', 'Rosters']
            
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                
                # Check if this sheet should have filters
                is_filterable = any(filterable in sheet_name for filterable in filterable_sheets)
                is_matchup = 'Week_' in sheet_name and 'Matchups' in sheet_name
                
                if is_filterable or is_matchup:
                    # Only add filters if there's data
                    if worksheet.max_row > 1 and worksheet.max_column > 0:
                        # Add auto-filter
                        filter_range = f"A1:{worksheet.cell(worksheet.max_row, worksheet.max_column).coordinate}"
                        worksheet.auto_filter.ref = filter_range
                        
                        # Freeze the top row (headers)
                        if freeze_panes:
                            worksheet.freeze_panes = "A2"
                        
                        # Make headers bold
                        if bold_headers:
                            for cell in worksheet[1]:  # First row
                                cell.font = Font(bold=True)
                        
                        print(f"  ✓ Enhanced '{sheet_name}' - filters, freeze panes, bold headers")
                    else:
                        print(f"  ⚠️ Skipped '{sheet_name}' - no data")
                else:
                    print(f"  ⏭️ Skipped '{sheet_name}' - not a data sheet")
            
            # Save the workbook
            workbook.save(filename)
            print("✓ Excel enhancements completed")
        
        except Exception as e:
            print(f"⚠️ Could not enhance Excel file: {e}")

    def auto_resize_columns(self, filename):
        """Auto-resize all columns in the Excel file to fit content"""
        try:
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter
            
            print("🔧 Auto-resizing columns...")
            
            # Load the workbook
            workbook = load_workbook(filename)
            
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]

                # Check if the sheet has filters, for padding
                has_filters = worksheet.auto_filter.ref is not None
                extra_padding = 3 if has_filters else 0
                
                # Auto-resize each column
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    
                    for cell in column:
                        try:
                            # Get the length of the cell value
                            if cell.value:
                                cell_length = len(str(cell.value))
                                if cell_length > max_length:
                                    max_length = cell_length
                        except:
                            pass
                    
                    # Set column width (add some padding)
                    total_padding = 2 + extra_padding
                    adjusted_width = min(max_length + total_padding, 50)  # Cap at 50 characters
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Save the workbook
            workbook.save(filename)
            print("✓ Column auto-resizing completed")
        
        except Exception as e:
            print(f"⚠️ Could not auto-resize columns: {e}")

    def safe_numeric(self, value, default=None):
        """Safely convert value to numeric, handling None and invalid values"""
        if value is None:
            return default
        
        # If it's already a number, return it
        if isinstance(value, (int, float)):
            return value
        
        # Try to convert string to number
        try:
            # Remove any whitespace
            str_value = str(value).strip()
            
            # Handle empty strings
            if not str_value:
                return default
            
            # Try integer first
            if '.' not in str_value:
                return int(str_value)
            else:
                return float(str_value)
        except (ValueError, TypeError):
            return default

    def parse_height(self, height_value):
        """Parse height from various formats to inches (numeric)"""
        if height_value is None:
            return None
        
        height_str = str(height_value).strip()
        if not height_str:
            return None
        
        try:
            # If it's already a number (inches), return it
            if height_str.isdigit():
                return int(height_str)
            
            # Handle feet'inches format like "6'2" or "6'2""
            if "'" in height_str:
                # Remove quotes and extra characters
                height_str = height_str.replace('"', '').replace("'", "'")
                
                # Split on apostrophe
                parts = height_str.split("'")
                if len(parts) == 2:
                    feet = int(parts[0])
                    inches = int(parts[1]) if parts[1] else 0
                    return (feet * 12) + inches
            
            # Handle dash format like "6-2"
            if "-" in height_str:
                parts = height_str.split("-")
                if len(parts) == 2:
                    feet = int(parts[0])
                    inches = int(parts[1]) if parts[1] else 0
                    return (feet * 12) + inches
            
            # Try to parse as float (already in inches)
            return float(height_str)
            
        except (ValueError, IndexError):
            return None

    def fuzzy_match_players(self, ranking_name, players_df, threshold=0.8):
        """Use fuzzy matching for names that don't match exactly"""
        try:
            from difflib import SequenceMatcher
        except ImportError:
            return None
        
        normalized_ranking = self.normalize_player_name(ranking_name)
        best_match = None
        best_score = 0
        
        for idx, row in players_df.iterrows():
            player_name = self.normalize_player_name(row['full_name'])
            
            # Calculate similarity
            similarity = SequenceMatcher(None, normalized_ranking, player_name).ratio()
            
            if similarity > best_score and similarity >= threshold:
                best_score = similarity
                best_match = idx
        
        return best_match, best_score if best_match is not None else (None, 0)

    def normalize_player_name(self, name):
        """Normalize player names for better matching"""
        if not name:
            return ""
        
        import unicodedata
        
        # Convert to lowercase and remove extra spaces
        normalized = str(name).lower().strip()
        
        # Handle accented characters - normalize to ASCII
        normalized = unicodedata.normalize('NFD', normalized)
        normalized = ''.join(char for char in normalized if unicodedata.category(char) != 'Mn')
        
        # Remove common suffixes
        suffixes = [' jr.', ' jr', ' sr.', ' sr', ' iii', ' ii', ' iv', ' v']
        for suffix in suffixes:
            if normalized.endswith(suffix):
                normalized = normalized[:-len(suffix)].strip()
        
        # Remove periods, hyphens, apostrophes, and other punctuation
        normalized = normalized.replace('.', '').replace('-', ' ').replace("'", '').replace('`', '')
        
        # Handle common name variations
        name_replacements = {
            'kenneth': 'ken',
            'michael': 'mike',
            'robert': 'bob',
            'william': 'will',
            'christopher': 'chris',
            'matthew': 'matt',
            'anthony': 'tony',
            'joshua': 'josh',
            'marquise': 'hollywood'  # Hollywood Brown is often listed as Marquise Brown
        }
        
        for full_name, short_name in name_replacements.items():
            normalized = normalized.replace(full_name, short_name)
        
        # Clean up multiple spaces
        normalized = ' '.join(normalized.split())
        
        return normalized

    def load_ranking_csvs(self, ffpc_csv, underdog_csv):
        """Load FFPC and Underdog CSV files and return DataFrames"""
        ranking_data = {}

        print(f"Loading FFPC rankings from {ffpc_csv}...")
        try:
            ffpc_df = pd.read_csv(ffpc_csv)
            # Standardize column names
            ffpc_df = ffpc_df.rename(columns={
                'Name': 'name',
                'Position': 'position', 
                'Team': 'team',
                'ADP': 'ffpc_adp',
                'ETR_Rank': 'ffpc_etr_rank',
                'Delta': 'ffpc_delta',
                'Pos_Rank': 'ffpc_pos_rank'
            })
            ranking_data['ffpc'] = ffpc_df
            print(f"✓ Loaded {len(ffpc_df)} FFPC rankings")
        except Exception as e:
            print(f"❌ Error loading FFPC CSV: {e}")
            sys.exit(1)

        print(f"Loading Underdog rankings from {underdog_csv}...")
        try:
            underdog_df = pd.read_csv(underdog_csv)
            # Standardize column names
            underdog_df = underdog_df.rename(columns={
                'Name': 'name',
                'Pos': 'position',
                'Team': 'team',
                'ADP': 'ud_adp',
                'ETR_Rank': 'ud_etr_rank',
                'Delta': 'ud_delta',
                'Pos_Rank': 'ud_pos_rank'
            })
                
            ranking_data['underdog'] = underdog_df
            print(f"✓ Loaded {len(underdog_df)} Underdog rankings")
        except Exception as e:
            print(f"❌ Error loading Underdog CSV: {e}")
            sys.exit(1)

        return ranking_data

    def merge_ranking_data(self, players_df, ranking_data):
        """Merge ranking data with players DataFrame"""
        if players_df.empty or not ranking_data:
            print("⚠️ No players data or ranking data to merge")
            return players_df, {}
        
        print("🔗 Merging ranking data with players...")
        
        # Add normalized name to players df for matching
        players_df['normalized_name'] = players_df['full_name'].apply(self.normalize_player_name)
        
        # Track matches and misses
        match_stats = {
            'ffpc_matched': 0,
            'ffpc_fuzzy_matched': 0,
            'ffpc_unmatched': [],
            'underdog_matched': 0,
            'underdog_fuzzy_matched': 0, 
            'underdog_unmatched': [],
            'players_with_rankings': 0
        }
        
        # Initialize ranking columns
        ranking_columns = [
            'ffpc_adp', 'ffpc_etr_rank', 'ffpc_delta', 'ffpc_pos_rank',
            'ud_adp', 'ud_etr_rank', 'ud_delta', 'ud_pos_rank'
        ]
        
        for col in ranking_columns:
            players_df[col] = None
        
        # Merge FFPC data
        if 'ffpc' in ranking_data:
            ffpc_df = ranking_data['ffpc']
            
            for idx, ffpc_row in ffpc_df.iterrows():
                # Normalize the ranking name for comparison
                normalized_ranking_name = self.normalize_player_name(ffpc_row['name'])
                
                # Try exact match first
                matches = players_df[players_df['normalized_name'] == normalized_ranking_name]
                
                if len(matches) == 1:
                    # Found exact match
                    player_idx = matches.index[0]
                    players_df.loc[player_idx, 'ffpc_adp'] = ffpc_row['ffpc_adp']
                    players_df.loc[player_idx, 'ffpc_etr_rank'] = ffpc_row['ffpc_etr_rank'] 
                    players_df.loc[player_idx, 'ffpc_delta'] = ffpc_row['ffpc_delta']
                    players_df.loc[player_idx, 'ffpc_pos_rank'] = ffpc_row['ffpc_pos_rank']
                    
                    match_stats['ffpc_matched'] += 1
                elif len(matches) > 1:
                    # Multiple direct matches
                    print(f"  ⚠️ Warning: Multiple exact matches for '{ffpc_row['name']}' - {len(matches)} found")
                    # Compare ffpc team and position to disambiguate
                    for player_idx in matches.index:
                        player_row = players_df.loc[player_idx]
                        if (player_row['team'] == ffpc_row['team']):
                            # Found the best match
                            players_df.loc[player_idx, 'ffpc_adp'] = ffpc_row['ffpc_adp']
                            players_df.loc[player_idx, 'ffpc_etr_rank'] = ffpc_row['ffpc_etr_rank'] 
                            players_df.loc[player_idx, 'ffpc_delta'] = ffpc_row['ffpc_delta']
                            players_df.loc[player_idx, 'ffpc_pos_rank'] = ffpc_row['ffpc_pos_rank']
                            
                            match_stats['ffpc_matched'] += 1
                            break
                    else:
                        # No disambiguation found
                        match_stats['ffpc_unmatched'].append({
                            'name': ffpc_row['name'],
                            'normalized_name': normalized_ranking_name,
                            'team': ffpc_row['team'],
                            'position': ffpc_row['position']
                        })
                else:
                    # Try fuzzy matching
                    fuzzy_match, similarity_score = self.fuzzy_match_players(ffpc_row['name'], players_df)
                    
                    if fuzzy_match is not None:
                        # Found fuzzy match
                        players_df.loc[fuzzy_match, 'ffpc_adp'] = ffpc_row['ffpc_adp']
                        players_df.loc[fuzzy_match, 'ffpc_etr_rank'] = ffpc_row['ffpc_etr_rank'] 
                        players_df.loc[fuzzy_match, 'ffpc_delta'] = ffpc_row['ffpc_delta']
                        players_df.loc[fuzzy_match, 'ffpc_pos_rank'] = ffpc_row['ffpc_pos_rank']
                        
                        match_stats['ffpc_fuzzy_matched'] += 1
                        print(f"  🎯 Fuzzy matched '{ffpc_row['name']}' → '{players_df.loc[fuzzy_match, 'full_name']}' (score: {similarity_score:.2f})")
                    else:
                        # No match found
                        match_stats['ffpc_unmatched'].append({
                            'name': ffpc_row['name'],
                            'normalized_name': normalized_ranking_name,
                            'team': ffpc_row['team'],
                            'position': ffpc_row['position']
                        })
        
        # Similar logic for Underdog data...
        if 'underdog' in ranking_data:
            underdog_df = ranking_data['underdog']
            
            for idx, underdog_row in underdog_df.iterrows():
                normalized_ranking_name = self.normalize_player_name(underdog_row['name'])
                matches = players_df[players_df['normalized_name'] == normalized_ranking_name]
                
                if len(matches) > 0:
                    # Exact match
                    player_idx = matches.index[0]
                    players_df.loc[player_idx, 'ud_adp'] = underdog_row['ud_adp']
                    players_df.loc[player_idx, 'ud_etr_rank'] = underdog_row['ud_etr_rank']
                    players_df.loc[player_idx, 'ud_delta'] = underdog_row['ud_delta']
                    players_df.loc[player_idx, 'ud_pos_rank'] = underdog_row['ud_pos_rank']
                    
                    match_stats['underdog_matched'] += 1
                else:
                    # Try fuzzy matching
                    fuzzy_match, similarity_score = self.fuzzy_match_players(underdog_row['name'], players_df)
                    
                    if fuzzy_match is not None:
                        players_df.loc[fuzzy_match, 'ud_adp'] = underdog_row['ud_adp']
                        players_df.loc[fuzzy_match, 'ud_etr_rank'] = underdog_row['ud_etr_rank']
                        players_df.loc[fuzzy_match, 'ud_delta'] = underdog_row['ud_delta']
                        players_df.loc[fuzzy_match, 'ud_pos_rank'] = underdog_row['ud_pos_rank']
                        
                        match_stats['underdog_fuzzy_matched'] += 1
                        print(f"  🎯 Fuzzy matched '{underdog_row['name']}' → '{players_df.loc[fuzzy_match, 'full_name']}' (score: {similarity_score:.2f})")
                    else:
                        match_stats['underdog_unmatched'].append({
                            'name': underdog_row['name'],
                            'normalized_name': normalized_ranking_name,
                            'team': underdog_row['team'],
                            'position': underdog_row['position'],
                        })
        
        # Count players with any ranking data
        has_rankings = (
            players_df['ffpc_adp'].notna() | 
            players_df['ud_adp'].notna()
        )
        match_stats['players_with_rankings'] = has_rankings.sum()
        
        # Remove the temporary normalized_name column
        players_df = players_df.drop('normalized_name', axis=1)
        
        # Print enhanced merge statistics
        print(f"\n🔗 Ranking merge results:")
        if 'ffpc' in ranking_data:
            ffpc_total = len(ranking_data['ffpc'])
            ffpc_exact = match_stats['ffpc_matched']
            ffpc_fuzzy = match_stats['ffpc_fuzzy_matched']
            ffpc_combined = ffpc_exact + ffpc_fuzzy
            print(f"  📊 FFPC: {ffpc_combined}/{ffpc_total} matched ({(ffpc_combined/ffpc_total*100):.1f}%) - {ffpc_exact} exact, {ffpc_fuzzy} fuzzy")
        
        if 'underdog' in ranking_data:
            underdog_total = len(ranking_data['underdog'])
            underdog_exact = match_stats['underdog_matched']
            underdog_fuzzy = match_stats['underdog_fuzzy_matched']
            underdog_combined = underdog_exact + underdog_fuzzy
            print(f"  📊 Underdog: {underdog_combined}/{underdog_total} matched ({(underdog_combined/underdog_total*100):.1f}%) - {underdog_exact} exact, {underdog_fuzzy} fuzzy")
        
        print(f"  👥 Players with rankings: {match_stats['players_with_rankings']}/{len(players_df)}")
        
        return players_df, match_stats

    def export_unmatched_rankings(self, match_stats, filename_base):
        """Export unmatched ranking players to CSV for review"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Export FFPC unmatched
        if match_stats['ffpc_unmatched']:
            ffpc_unmatched_file = f"{filename_base}_ffpc_unmatched_{timestamp}.csv"
            ffpc_unmatched_df = pd.DataFrame(match_stats['ffpc_unmatched'])
            ffpc_unmatched_df.to_csv(ffpc_unmatched_file, index=False)
            print(f"📄 Exported {len(match_stats['ffpc_unmatched'])} unmatched FFPC players to: {ffpc_unmatched_file}")
        
        # Export Underdog unmatched  
        if match_stats['underdog_unmatched']:
            underdog_unmatched_file = f"{filename_base}_underdog_unmatched_{timestamp}.csv"
            underdog_unmatched_df = pd.DataFrame(match_stats['underdog_unmatched'])
            underdog_unmatched_df.to_csv(underdog_unmatched_file, index=False)
            print(f"📄 Exported {len(match_stats['underdog_unmatched'])} unmatched Underdog players to: {underdog_unmatched_file}")

    def add_position_conditional_formatting_separate(self, filename):
        """Add conditional formatting with separate rules for better reliability"""
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill
            from openpyxl.formatting.rule import FormulaRule
            from openpyxl.utils import get_column_letter
            
            print("🎨 Adding position conditional formatting...")
            
            position_colors = {
                'QB': 'FFCCCB', 'RB': 'C8E6C9', 'WR': 'BBDEFB', 
                'TE': 'FFE0B2', 'K': 'E1BEE7', 'DEF': 'EEEEEE', 'DST': 'EEEEEE'
            }
            
            workbook = load_workbook(filename)
            
            if 'Players' not in workbook.sheetnames:
                print("  ⚠️ No Players sheet found")
                return False
                
            worksheet = workbook['Players']
            
            # Find position column
            position_col = None
            for col_num, cell in enumerate(worksheet[1], 1):
                if cell.value and 'position' in str(cell.value).lower():
                    position_col = col_num
                    break
            
            if not position_col:
                print("  ⚠️ Position column not found")
                return False
            
            position_col_letter = get_column_letter(position_col)
            
            # Get data dimensions
            last_row = worksheet.max_row
            last_col_letter = get_column_letter(worksheet.max_column)
            
            print(f"  📍 Data range: A2:{last_col_letter}{last_row}")
            print(f"  🔍 Position column: {position_col_letter}")
            
            # Apply formatting for each position separately
            for position, color_hex in position_colors.items():
                try:
                    # Create fill
                    fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid')
                    
                    # Create range for this rule
                    range_ref = f"A2:{last_col_letter}{last_row}"
                    
                    # Formula: check if position column equals this position
                    # Use absolute column reference, relative row reference
                    formula = f'${position_col_letter}2="{position}"'
                    
                    # Create and add rule
                    rule = FormulaRule(formula=[formula], fill=fill)
                    worksheet.conditional_formatting.add(range_ref, rule)
                    
                    print(f"    ✓ {position}: {formula} -> #{color_hex}")
                    
                except Exception as e:
                    print(f"    ❌ Failed {position}: {e}")
            
            workbook.save(filename)
            workbook.close()
            
            print("✓ Position formatting applied")
            return True
            
        except Exception as e:
            print(f"⚠️ Formatting error: {e}")
            return False

    def export_to_excel(self, filename=None, league_id=None, include_players=False, 
                   weeks=None, use_cache=True, ffpc_csv=None, underdog_csv=None):
        """Export data to Excel"""
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"sleeper_data_{timestamp}.xlsx"
        
        # Get all data (cached or fresh)
        all_data = self.get_data(league_id, include_players, weeks, use_cache)
        
        if not all_data:
            print("❌ No data to export")
            return None
        
        cache_status = "from cache" if use_cache and self.cache_file.exists() else "fresh from API"
        print(f"Starting export to {filename} ({cache_status})...")
        
        # Debug: Show what data we have
        print(f"🔍 Debug - Available data keys: {list(all_data.keys())}")
        
        sheets_created = 0
        
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            
            # Export players if available (with smart filtering)
            if 'players' in all_data and all_data['players']:
                print(f"🔍 Debug - Players data type: {type(all_data['players'])}, length: {len(all_data['players']) if all_data['players'] else 0}")
                
                # Pass league data for filtering
                league_data = all_data.get('league_info')
                players_df = self.process_players_data(all_data['players'], league_data)

                print(f" 🔍 Printing players data to CSV...")
                try:
                    players_df.to_csv('players_data.csv', index=False)
                    print("✓ Players data exported to players_data.csv")
                except Exception as e:
                    print(f"❌ Failed to export players data to CSV: {e}")
                
                if not players_df.empty and (ffpc_csv and underdog_csv):
                    try:
                        ranking_data = self.load_ranking_csvs(ffpc_csv, underdog_csv)
                        players_df, match_stats = self.merge_ranking_data(players_df, ranking_data)

                        # Export unmatched rankings for review
                        base_name = filename.replace('.xlsx', '')
                        self.export_unmatched_rankings(match_stats, base_name)
                    except Exception as e:
                        print(f"⚠️ Warning: Ranking merge failed ({e}), continuing with original player data")

                if not players_df.empty:
                    try:
                        cols_to_remove = ['first_name', 'last_name', 'active']
                        players_df = players_df.drop(columns=cols_to_remove, errors='ignore')
                        players_df.to_excel(writer, sheet_name='Players', index=False)
                        print(f"✓ Exported {len(players_df)} fantasy-relevant players")
                        sheets_created += 1
                    except Exception as e:
                        print(f"❌ Failed to export players sheet: {e}")
                        # Try exporting without rankings
                        try:
                            league_data = all_data.get('league_info')
                            clean_players_df = self.process_players_data(all_data['players'], league_data)
                            clean_players_df.to_excel(writer, sheet_name='Players', index=False)
                            print(f"✓ Exported {len(clean_players_df)} players (without rankings)")
                            sheets_created += 1
                        except Exception as e2:
                            print(f"❌ Failed to export players entirely: {e2}")
                else:
                    print("⚠️ Players DataFrame is empty after filtering")
            
            # Export league data if available
            if league_id:
                # League info
                if 'league_info' in all_data and all_data['league_info']:
                    league_df = pd.DataFrame([all_data['league_info']])
                    league_df.to_excel(writer, sheet_name='League_Info', index=False)
                    print("✓ Exported league info")
                    sheets_created += 1
                    
                    # Show roster positions for reference
                    if 'roster_positions' in all_data['league_info']:
                        print(f"🔍 League roster positions: {all_data['league_info']['roster_positions']}")
                else:
                    print("⚠️ No league info data found")
                
                # Users
                if 'users' in all_data and all_data['users']:
                    pd.DataFrame(all_data['users']).to_excel(writer, sheet_name='Users', index=False)
                    print(f"✓ Exported {len(all_data['users'])} users")
                    sheets_created += 1
                else:
                    print("⚠️ No users data found")
                
                # Rosters
                if 'rosters' in all_data and all_data['rosters']:
                    rosters_df = self.process_rosters_data(all_data['rosters'], all_data.get('users'))
                    if not rosters_df.empty:
                        rosters_df.to_excel(writer, sheet_name='Rosters', index=False)
                        print(f"✓ Exported {len(rosters_df)} rosters")
                        sheets_created += 1
                    else:
                        print("⚠️ Rosters DataFrame is empty")
                else:
                    print("⚠️ No rosters data found")
                
                # Matchups
                if 'matchups' in all_data and all_data['matchups']:
                    for week, matchup_data in all_data['matchups'].items():
                        if matchup_data:
                            matchups_df = pd.DataFrame(matchup_data)
                            matchups_df.to_excel(writer, sheet_name=f'Week_{week}_Matchups', index=False)
                            print(f"✓ Exported week {week} matchups")
                            sheets_created += 1
            
            # Always export metadata as a fallback
            if '_metadata' in all_data:
                metadata_df = pd.DataFrame([all_data['_metadata']])
                metadata_df.to_excel(writer, sheet_name='Metadata', index=False)
                print("✓ Exported metadata")
                sheets_created += 1
            
            # Emergency fallback - create a summary sheet if no other sheets were created
            if sheets_created == 0:
                print("⚠️ No sheets created, adding summary sheet as fallback")
                summary_data = {
                    'Data Available': list(all_data.keys()),
                    'Status': ['Available' if all_data[key] else 'Empty' for key in all_data.keys()]
                }
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name='Summary', index=False)
                sheets_created += 1
        
        # Add filters and formatting to the Excel file
        self.add_excel_filters(filename)

        # Auto-resize columns for better readability
        self.auto_resize_columns(filename)

        # Add position-based conditional formatting
        self.add_position_conditional_formatting_separate(filename)

        print(f"🎉 Export completed: {filename} ({sheets_created} sheets created)")
        return filename

    def clear_cache(self):
        """Clear the cache file"""
        if self.cache_file.exists():
            try:
                self.cache_file.unlink()
                print("🗑️ Cache cleared")
            except OSError as e:
                print(f"❌ Failed to clear cache: {e}")
        else:
            print("No cache file to clear")
    
    def show_cache_info(self):
        """Show cache file info"""
        if not self.cache_file.exists():
            print("No cache file found")
            return
        
        # File info
        size_mb = self.cache_file.stat().st_size / (1024 * 1024)
        modified = datetime.fromtimestamp(self.cache_file.stat().st_mtime)
        
        print(f"\n=== Cache Info ===")
        print(f"File: {self.cache_file}")
        print(f"Size: {size_mb:.2f} MB")
        print(f"Modified: {modified.strftime('%Y-%m-%d %H:%M:%S')}")
        
        # Try to load and show metadata
        try:
            with open(self.cache_file, 'rb') as f:
                data = pickle.load(f)
            
            metadata = data.get('_metadata', {})
            if metadata:
                print(f"Fetched: {metadata.get('fetched_at', 'Unknown')}")
                print(f"League ID: {metadata.get('league_id', 'None')}")
                print(f"Includes players: {metadata.get('included_players', False)}")
                print(f"Weeks: {metadata.get('weeks', [])}")
        except:
            print("Could not read cache metadata")

def main():
    """Example usage with league ID validation and cache fallback"""
    exporter = SleeperAPIExporter()
    
    # Show current cache
    exporter.show_cache_info()
    
    # Test different league IDs - replace these with real ones to test
    test_league_ids = [
        "1235124102474235904"
    ]

    USE_CACHE = True  # Set to False to force fresh API calls
    
    print("\n=== Testing League ID Validation ===")
    
    valid_league_id = None
    
    for league_id in test_league_ids:
        print(f"\nTesting league ID: {league_id}")
        
        try:
            if exporter.validate_league_id(league_id):
                print(f"✅ League ID {league_id} is VALID!")
                valid_league_id = league_id
            else:
                print(f"❌ League ID {league_id} is invalid, skipping...")
                continue

        except Exception as e:
            print(f"❌ Error validating league ID {league_id}: {e}")
            continue
    
        if valid_league_id:
            print(f"\n=== Using Valid League ID: {valid_league_id} ===")
            
            if USE_CACHE:
                print("\n💾 --- Attempting to source data from pickle cache ---")
            else:
                print("\n🌐 --- Attempting to fetch fresh data from API ---")

            try:
                exporter.export_to_excel(
                    f"{valid_league_id}.xlsx", 
                    league_id=valid_league_id,
                    include_players=True,
                    use_cache=USE_CACHE,
                    ffpc_csv=os.path.join(os.path.dirname(__file__), "etr_rankings", "2025 FFPC Best Ball SuperFlex Rankings.csv"),
                    underdog_csv=os.path.join(os.path.dirname(__file__), "etr_rankings", "2025 Underdog SuperFlex Rankings.csv")
                )
            except Exception as e:
                print(f"❌ Export failed: {e}")
            
            # Show final cache info
            exporter.show_cache_info()
            
        else:
            print("\n❌ No valid league IDs found!  Source your league ID(s) from the Sleeper URL.  Exiting...")
            return  # Stop further processing

if __name__ == "__main__":
    main()
